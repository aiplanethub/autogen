"""
File message types for handling various file formats in AutoGen.

This module extends AutoGen's message system with specialized message types
for handling PDFs, JSON files, images, OCR-processed PDFs, and Excel sheets.
"""

import json
import os
from typing import Any, Dict, List, Literal, Optional, Union, Mapping

from autogen_agentchat.messages import BaseChatMessage, MessageFactory, BaseAgentEvent
from autogen_core import Image
from pydantic import BaseModel, Field, computed_field


class FileMessage(BaseChatMessage):
    """Base message type for file-based content."""

    filepath: str
    """Path to the file."""

    filename: str
    """Name of the file."""

    filetype: str
    """MIME type or extension of the file."""

    exists: bool = True
    """Whether the file exists (default: True)."""

    raw_content: Optional[Any] = Field(default=None, exclude=True)
    """Internal storage for file content."""

    metadata: Dict[str, str] = {}
    """Additional metadata about the file."""

    type: Literal["FileMessage"] = "FileMessage"

    def __init__(self, **data):
        super().__init__(**data)
        if self.exists and not os.path.exists(self.filepath):
            raise FileNotFoundError(f"File not found: {self.filepath}")
        if not self.filename:
            self.filename = os.path.basename(self.filepath)

    @computed_field
    def content(self) -> str:
        """Standard content field that all message types have."""
        return self.to_text()

    def to_text(self) -> str:
        return f"File: {self.filename} ({self.filetype})"

    def to_model_text(self) -> str:
        return f"File: {self.filename} ({self.filetype})"

    def to_model_message(self):
        from autogen_core.models import UserMessage

        return UserMessage(content=self.to_model_text(), source=self.source)
    
    async def get_content(self) -> Any:
        """Get the content of the file."""
        # Base implementation returns None, override in subclasses
        return None

class FileProcessingEvent(BaseAgentEvent):
    """An event signaling file processing activities."""

    filename: str
    """The name of the file being processed."""

    operation: str
    """The type of operation being performed (e.g., 'extracting', 'ocr', 'loading')."""

    status: Literal["started", "in_progress", "completed", "failed"] = "in_progress"
    """The current status of the file processing operation."""

    progress: float | None = None
    """Optional progress indicator (0.0 to 1.0) for long-running operations."""

    content: str | None = None
    """Additional details about the processing operation."""

    type: Literal["FileProcessingEvent"] = "FileProcessingEvent"

    def to_text(self) -> str:
        """Convert the event to a human-readable string."""
        status_text = {
            "started": "Starting",
            "in_progress": "Processing",
            "completed": "Completed processing",
            "failed": "Failed processing"
        }[self.status]
        
        message = f"{status_text} {self.operation} for file: {self.filename}"
        if self.progress is not None:
            message += f" ({int(self.progress * 100)}%)"
        if self.details:
            message += f"\nDetails: {self.details}"
        return message

    
class PDFMessage(FileMessage):
    """Message type for PDF files."""

    page_count: Optional[int] = None
    """Number of pages in the PDF."""

    title: Optional[str] = None
    """Title of the PDF document."""

    author: Optional[str] = None
    """Author of the PDF document."""

    extracted_text: Optional[str] = None
    """Extracted text content from the PDF."""

    type: Literal["PDFMessage"] = "PDFMessage"

    def __init__(self, **data):
        super().__init__(**data)
        if self.exists and self.filetype.lower() not in ["pdf", "application/pdf"]:
            raise ValueError("File is not a PDF")

        # Automatically extract metadata if file exists and library is available
        if self.exists:
            self._extract_metadata()

            # Set up async extraction for the next event loop iteration
            if self.extracted_text is None:
                import asyncio

                try:
                    loop = asyncio.get_event_loop()
                    if loop.is_running():
                        # Schedule extraction for later if we're in an event loop
                        loop.create_task(self._async_extract_text())
                    else:
                        # Extract synchronously if no event loop is running
                        self.extracted_text = self._extract_text_sync()
                except RuntimeError:
                    # No event loop, extract synchronously
                    self.extracted_text = self._extract_text_sync()

    @computed_field
    def content(self) -> str:
        """Standard content field that returns the extracted text."""
        return self.extracted_text if self.extracted_text else self.to_text()

    def _extract_metadata(self):
        """Extract metadata from the PDF file."""
        if not self.exists or self.page_count is not None:
            return

        try:
            import PyPDF2

            with open(self.filepath, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                self.page_count = len(reader.pages)
                if reader.metadata:
                    if self.title is None and reader.metadata.title:
                        self.title = reader.metadata.title
                    if self.author is None and reader.metadata.author:
                        self.author = reader.metadata.author
        except ImportError:
            pass  # PyPDF2 not installed
        except Exception:
            pass  # Error reading PDF

    def _extract_text_sync(self) -> str:
        """Extract text content from the PDF synchronously."""
        if not self.exists:
            return ""

        try:
            import PyPDF2

            text = []
            with open(self.filepath, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text.append(page.extract_text())
            return "\n\n".join(text)
        except ImportError:
            return "PyPDF2 library not installed. Cannot extract text."
        except Exception as e:
            return f"Error extracting text: {str(e)}"

    async def _async_extract_text(self):
        """Extract text content from the PDF asynchronously."""
        self.extracted_text = self._extract_text_sync()

    def to_text(self) -> str:
        base = super().to_text()
        details = []
        if self.page_count:
            details.append(f"{self.page_count} pages")
        if self.title:
            details.append(f"Title: {self.title}")
        if self.author:
            details.append(f"Author: {self.author}")

        if details:
            result = f"{base} ({', '.join(details)})"
        else:
            result = base

        # Include a snippet of extracted text if available
        if self.extracted_text:
            preview = self.extracted_text[:200] + "..." if len(self.extracted_text) > 200 else self.extracted_text
            result += f"\n\nContent preview:\n{preview}"

        return result

    def to_model_text(self) -> str:
        base_text = super().to_text()

        if self.extracted_text:
            return f"{base_text}\n\nContent:\n{self.extracted_text}"
        return base_text

    async def extract_text(self) -> str:
        """Extract text content from the PDF if not already extracted."""
        if self.extracted_text is None:
            self.extracted_text = self._extract_text_sync()
        return self.extracted_text
    
    async def get_content(self) -> str:
        """Get the content of the PDF file."""
        if self.extracted_text is None:
            await self.extract_text()
        return self.extracted_text


class PDFWithOCRMessage(PDFMessage):
    """Message type for PDF files with OCR processing."""

    ocr_applied: bool = False
    """Whether OCR has been applied to the PDF."""

    ocr_language: Optional[str] = None
    """Language used for OCR processing."""

    ocr_text: Optional[str] = None
    """Text extracted via OCR."""

    type: Literal["PDFWithOCRMessage"] = "PDFWithOCRMessage"

    def __init__(self, **data):
        super().__init__(**data)

        # If OCR language is specified but OCR hasn't been applied yet,
        # schedule OCR extraction for the next event loop iteration
        if self.exists and not self.ocr_applied and self.ocr_language and self.ocr_text is None:
            import asyncio

            try:
                loop = asyncio.get_event_loop()
                if loop.is_running():
                    # Schedule extraction for later if we're in an event loop
                    loop.create_task(self._async_apply_ocr())
                else:
                    # Extract synchronously if no event loop is running
                    self.ocr_text = self._apply_ocr_sync()
                    self.ocr_applied = True
            except (RuntimeError, ImportError):
                # No event loop or missing dependencies
                pass

    @computed_field
    def content(self) -> str:
        """Standard content field that prioritizes OCR text if available."""
        if self.ocr_text and not self.ocr_text.startswith("Required") and not self.ocr_text.startswith("Error"):
            return self.ocr_text
        elif self.extracted_text:
            return self.extracted_text
        return self.to_text()

    async def _async_apply_ocr(self):
        """Apply OCR to the PDF asynchronously."""
        self.ocr_text = self._apply_ocr_sync()
        self.ocr_applied = True
        self.extracted_text = self.ocr_text  # Update the base class text as well

    def _apply_ocr_sync(self) -> str:
        """Apply OCR to the PDF synchronously."""
        if not self.exists:
            return ""

        try:
            # Check if pytesseract and pdf2image are available
            import pytesseract
            from pdf2image import convert_from_path

            # Convert PDF to images
            images = convert_from_path(self.filepath)
            text = []

            # Apply OCR to each image
            for img in images:
                if self.ocr_language:
                    img_text = pytesseract.image_to_string(img, lang=self.ocr_language)
                else:
                    img_text = pytesseract.image_to_string(img)
                text.append(img_text)

            return "\n\n".join(text)
        except ImportError:
            return "Required OCR libraries (pytesseract, pdf2image) not installed."
        except Exception as e:
            return f"Error applying OCR: {str(e)}"

    def to_text(self) -> str:
        base = super().to_text()
        if self.ocr_applied:
            return f"{base} (OCR applied)"
        return f"{base} (OCR ready)"

    def to_model_text(self) -> str:
        base_text = super().to_model_text()

        # Include OCR text if available
        if self.ocr_text and not self.ocr_text.startswith("Required") and not self.ocr_text.startswith("Error"):
            return f"{base_text}\n\nOCR Text:\n{self.ocr_text}"
        return base_text

    async def extract_text(self) -> str:
        """Extract text content from the PDF using OCR if needed."""
        # If text is already extracted, return it
        if self.extracted_text is not None:
            return self.extracted_text

        # If OCR has already been applied, use that text
        if self.ocr_applied and self.ocr_text is not None:
            self.extracted_text = self.ocr_text
            return self.ocr_text

        # Otherwise, apply OCR
        self.ocr_text = self._apply_ocr_sync()
        self.ocr_applied = True
        self.extracted_text = self.ocr_text

        return self.extracted_text
    
    async def get_content(self) -> str:
        """Get the content of the PDF file with OCR."""
        if self.ocr_applied and self.ocr_text:
            await self.extract_text()
        return self.extracted_text


class JSONMessage(FileMessage):
    """Message type for JSON files."""

    schema: Optional[Dict[str, Any]] = None
    """JSON schema of the file content."""

    content_data: Optional[Any] = None
    """Parsed JSON content."""

    type: Literal["JSONMessage"] = "JSONMessage"

    def __init__(self, **data):
        super().__init__(**data)
        if self.exists and self.filetype.lower() not in ["json", "application/json"]:
            raise ValueError("File is not a JSON file")

        # Automatically load content if file exists
        if self.exists and self.content_data is None:
            try:
                with open(self.filepath, "r", encoding="utf-8") as f:
                    self.content_data = json.load(f)

                # Try to infer schema from the content
                self._infer_schema()
            except Exception:
                # Don't raise exception on initialization, defer to get_content
                pass

    @computed_field
    def content(self) -> Any:
        """Standard content field that returns JSON data."""
        return self.content_data

    def _infer_schema(self):
        """Infer a basic schema from the content."""
        if self.content_data is None or self.schema is not None:
            return

        try:
            def get_type(value):
                if isinstance(value, dict):
                    properties = {}
                    for k, v in value.items():
                        properties[k] = get_type(v)
                    return {"type": "object", "properties": properties}
                elif isinstance(value, list):
                    if value:
                        return {"type": "array", "items": get_type(value[0])}
                    return {"type": "array"}
                elif isinstance(value, str):
                    return {"type": "string"}
                elif isinstance(value, bool):
                    return {"type": "boolean"}
                elif isinstance(value, int):
                    return {"type": "integer"}
                elif isinstance(value, float):
                    return {"type": "number"}
                else:
                    return {"type": "null"}

            self.schema = get_type(self.content_data)
        except Exception:
            # Schema inference is a best-effort feature
            pass

    async def get_content(self) -> Any:
        """Parse and return the JSON content."""
        if self.content_data is not None:
            return self.content_data

        if not self.exists:
            return None

        try:
            with open(self.filepath, "r", encoding="utf-8") as f:
                self.content_data = json.load(f)
            return self.content_data
        except Exception as e:
            raise ValueError(f"Error parsing JSON file: {str(e)}")

    def to_text(self) -> str:
        base = f"JSON File: {self.filename}"

        if self.content_data is not None:
            # Add a preview of the JSON content
            preview = str(self.content_data)
            if len(preview) > 500:
                preview = preview[:500] + "..."
            return f"{base}\n\nContent preview:\n{preview}"
        return base

    def to_model_text(self) -> str:
        if self.content_data is not None:
            return f"JSON File: {self.filename}\n\nContent: {json.dumps(self.content_data, indent=2)}"
        return f"JSON File: {self.filename} (content not loaded yet)"


class ImageMessage(FileMessage):
    """Message type for image files."""

    width: Optional[int] = None
    """Width of the image in pixels."""

    height: Optional[int] = None
    """Height of the image in pixels."""

    ocr_text: Optional[str] = None
    """Text extracted from the image using OCR."""

    type: Literal["ImageMessage"] = "ImageMessage"

    def __init__(self, **data):
        super().__init__(**data)
        image_types = [
            "image/jpeg",
            "image/png",
            "image/gif",
            "image/webp",
            "image/tiff",
            "image/bmp",
            "image/svg+xml",
            "jpg",
            "jpeg",
            "png",
            "gif",
            "webp",
            "tiff",
            "bmp",
            "svg",
        ]
        if self.exists and self.filetype.lower() not in image_types:
            raise ValueError("File is not a supported image type")

        # Automatically extract image dimensions if file exists
        if self.exists:
            self._extract_dimensions()

            # Try to perform automatic OCR if possible
            import asyncio

            try:
                loop = asyncio.get_event_loop()
                if loop.is_running():
                    # Schedule OCR for later if we're in an event loop
                    loop.create_task(self._async_ocr_extract_text())
                else:
                    # Extract synchronously if no event loop is running
                    self.ocr_text = self._ocr_extract_text_sync()
            except (RuntimeError, ImportError):
                # No event loop or missing dependencies, skip automatic OCR
                pass

    @computed_field
    def content(self) -> Union[str, Dict[str, Any]]:
        """Standard content field that returns OCR text if available, otherwise image metadata."""
        if self.ocr_text and not self.ocr_text.startswith("Required") and not self.ocr_text.startswith("Error"):
            return self.ocr_text
        
        # Return image metadata as content if no OCR text
        content_dict = {
            "image": self.filename,
            "type": self.filetype
        }
        if self.width and self.height:
            content_dict["dimensions"] = f"{self.width}x{self.height}"
        
        return content_dict

    def _extract_dimensions(self):
        """Extract dimensions from the image file."""
        if not self.exists or (self.width is not None and self.height is not None):
            return

        try:
            from PIL import Image as PILImage

            with PILImage.open(self.filepath) as img:
                self.width, self.height = img.size
        except ImportError:
            pass  # PIL not installed
        except Exception:
            pass  # Error reading image

    def _ocr_extract_text_sync(self, language: Optional[str] = None) -> str:
        """Extract text from the image using OCR synchronously."""
        if not self.exists:
            return ""

        try:
            import pytesseract
            from PIL import Image as PILImage

            with PILImage.open(self.filepath) as img:
                if language:
                    text = pytesseract.image_to_string(img, lang=language)
                else:
                    text = pytesseract.image_to_string(img)
                return text
        except ImportError:
            return "Required OCR library (pytesseract) not installed."
        except Exception as e:
            return f"Error applying OCR: {str(e)}"

    async def _async_ocr_extract_text(self, language: Optional[str] = None):
        """Extract text from the image using OCR asynchronously."""
        self.ocr_text = self._ocr_extract_text_sync(language)

    def to_text(self) -> str:
        base = super().to_text()
        details = []

        if self.width and self.height:
            details.append(f"{self.width}x{self.height}")

        if details:
            result = f"{base} ({', '.join(details)})"
        else:
            result = base

        # Include extracted OCR text if available
        if (
            self.ocr_text
            and self.ocr_text.strip()
            and not self.ocr_text.startswith("Error")
            and not self.ocr_text.startswith("Required")
        ):
            preview = self.ocr_text.strip()
            if len(preview) > 200:
                preview = preview[:200] + "..."
            result += f"\n\nOCR Text:\n{preview}"

        return result

    def to_model_text(self) -> str:
        base = super().to_text()

        # Include image dimensions if available
        if self.width and self.height:
            base += f" ({self.width}x{self.height})"

        # Include OCR text if available
        if (
            self.ocr_text
            and self.ocr_text.strip()
            and not self.ocr_text.startswith("Error")
            and not self.ocr_text.startswith("Required")
        ):
            base += f"\n\nImage text content:\n{self.ocr_text.strip()}"

        return base

    def to_model_message(self):
        from autogen_core.models import UserMessage

        # If the file exists, create a multimodal message with the image
        if self.exists:
            try:
                # Convert to autogen_core.Image
                with open(self.filepath, "rb") as f:
                    file_content = f.read()

                image_obj = Image(
                    data=file_content, mime_type=self.filetype if "/" in self.filetype else f"image/{self.filetype}"
                )

                # For the content, include both the image and any OCR text if available
                if self.ocr_text and not self.ocr_text.startswith("Error") and not self.ocr_text.startswith("Required"):
                    return UserMessage(
                        content=[f"Image: {self.filename} with text content:\n{self.ocr_text.strip()}", image_obj],
                        source=self.source,
                    )
                else:
                    return UserMessage(content=[f"Image: {self.filename}", image_obj], source=self.source)
            except Exception:
                # Fall back to text if image loading fails
                return UserMessage(content=self.to_model_text(), source=self.source)
        else:
            return UserMessage(content=self.to_model_text(), source=self.source)

    async def ocr_extract_text(self, language: Optional[str] = None) -> str:
        """Extract text from the image using OCR."""
        if self.ocr_text is not None:
            return self.ocr_text

        self.ocr_text = self._ocr_extract_text_sync(language)
        return self.ocr_text
    
    async def get_content(self) -> Union[str, Dict[str, Any]]:
        """Get the content of the image file."""
        # If OCR text is available, return it as the content
        if self.ocr_text and not self.ocr_text.startswith("Required") and not self.ocr_text.startswith("Error"):
            return self.ocr_text
        
        # Otherwise, return the image metadata
        return self.content  # Use the computed field


class ExcelMessage(FileMessage):
    """Message type for Excel spreadsheet files."""

    sheet_names: Optional[List[str]] = None
    """Names of sheets in the workbook."""

    row_count: Optional[Dict[str, int]] = None
    """Number of rows per sheet."""

    column_count: Optional[Dict[str, int]] = None
    """Number of columns per sheet."""

    dataframes: Optional[Dict[str, Any]] = None
    """Cached pandas DataFrames from the Excel file."""

    data_dict: Optional[Dict[str, List[Dict[str, Any]]]] = None
    """Cached dictionary representation of the Excel data."""

    type: Literal["ExcelMessage"] = "ExcelMessage"

    def __init__(self, **data):
        super().__init__(**data)
        excel_types = [
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xls",
            "xlsx",
            "xlsm",
            "xlsb",
            "csv",
        ]
        if self.exists and self.filetype.lower() not in excel_types:
            raise ValueError("File is not a supported Excel format")

        # Automatically extract workbook info if file exists
        if self.exists and not self.sheet_names:
            self._extract_workbook_info()

            # Schedule data extraction for the next event loop iteration
            if self.dataframes is None:
                import asyncio

                try:
                    loop = asyncio.get_event_loop()
                    if loop.is_running():
                        # Schedule extraction for later if we're in an event loop
                        loop.create_task(self._async_extract_data())
                except (RuntimeError, ImportError):
                    # No event loop or missing dependencies, skip automatic extraction
                    pass

    @computed_field
    def content(self) -> Dict[str, Any]:
        """Standard content field that returns Excel data as a dictionary."""
        if self.data_dict:
            return self.data_dict
        
        # If data_dict is not available yet, return sheet info
        content_dict = {
            "file": self.filename,
            "type": self.filetype
        }
        
        if self.sheet_names:
            sheets_info = {}
            for sheet in self.sheet_names:
                sheet_info = {}
                if self.row_count and self.column_count:
                    sheet_info["rows"] = self.row_count.get(sheet, 0)
                    sheet_info["columns"] = self.column_count.get(sheet, 0)
                sheets_info[sheet] = sheet_info
            content_dict["sheets"] = sheets_info
            
        return content_dict

    def _extract_workbook_info(self):
        """Extract information about the workbook."""
        if not self.exists or self.sheet_names is not None:
            return

        # Try with pandas first
        try:
            import pandas as pd
            
            # Determine the engine based on file extension
            engine = None
            file_ext = os.path.splitext(self.filepath)[1].lower()
            
            if file_ext in ['.xlsx', '.xlsm']:
                engine = 'openpyxl'
            elif file_ext == '.xls':
                engine = 'xlrd'
            elif file_ext == '.xlsb':
                engine = 'pyxlsb'
            
            # For CSV, try using read_csv instead
            if file_ext == '.csv':
                df = pd.read_csv(self.filepath)
                self.sheet_names = ['Sheet1']  # CSV files have a single sheet
                self.row_count = {'Sheet1': len(df)}
                self.column_count = {'Sheet1': len(df.columns)}
                return
            
            # Try with the selected engine first
            try:
                xl = pd.ExcelFile(self.filepath, engine=engine)
            except Exception:
                # If initial attempt fails, try other engines
                for alt_engine in ['openpyxl', 'xlrd', 'pyxlsb']:
                    if alt_engine == engine:  # Skip already tried engine
                        continue
                    try:
                        xl = pd.ExcelFile(self.filepath, engine=alt_engine)
                        break  # Found a working engine
                    except Exception:
                        continue
            
            self.sheet_names = xl.sheet_names
            self.row_count = {}
            self.column_count = {}

            for sheet in self.sheet_names:
                df = xl.parse(sheet)
                self.row_count[sheet] = len(df)
                self.column_count[sheet] = len(df.columns)

            return
        except ImportError:
            pass  # pandas not installed
        except Exception:
            pass  # Error with pandas

        # Try with openpyxl as fallback
        try:
            import openpyxl

            wb = openpyxl.load_workbook(self.filepath, read_only=True)
            self.sheet_names = wb.sheetnames
            self.row_count = {}
            self.column_count = {}

            for sheet in self.sheet_names:
                ws = wb[sheet]
                self.row_count[sheet] = ws.max_row
                self.column_count[sheet] = ws.max_column
        except ImportError:
            pass  # openpyxl not installed
        except Exception:
            pass  # Error reading Excel file

    async def _async_extract_data(self):
        """Extract data from the Excel file asynchronously."""
        try:
            self.dataframes = await self.to_dataframes()
            self.data_dict = await self.to_dict()
        except Exception:
            # Silent failure for background extraction
            pass

    def to_text(self) -> str:
        base = super().to_text()
        if self.sheet_names:
            sheet_info = []
            for sheet in self.sheet_names:
                info = f"'{sheet}'"
                if self.row_count and self.column_count:
                    rows = self.row_count.get(sheet, "?")
                    cols = self.column_count.get(sheet, "?")
                    info += f" ({rows}x{cols})"
                sheet_info.append(info)

            result = f"{base} (Sheets: {', '.join(sheet_info)})"

            # Include data preview if available
            if self.data_dict:
                preview = str(self.data_dict)
                if len(preview) > 300:
                    preview = preview[:300] + "..."
                result += f"\n\nData preview:\n{preview}"

            return result
        return base

    def to_model_text(self) -> str:
        base_text = super().to_text()

        # Add sheet information
        if self.sheet_names:
            sheets_info = []
            for sheet in self.sheet_names:
                if self.row_count and self.column_count:
                    rows = self.row_count.get(sheet, "?")
                    cols = self.column_count.get(sheet, "?")
                    sheets_info.append(f"'{sheet}' ({rows}x{cols})")
                else:
                    sheets_info.append(f"'{sheet}'")
            base_text += f"\nSheets: {', '.join(sheets_info)}"

        # Add data if available
        if self.data_dict:
            data_str = json.dumps(self.data_dict, indent=2)
            if len(data_str) > 2000:  # Limit data size for model context
                data_str = data_str[:2000] + "...[truncated]"
            base_text += f"\n\nData:\n{data_str}"

        return base_text

    async def to_dataframes(self) -> Dict[str, Any]:
        """Convert Excel file to pandas DataFrames."""
        if self.dataframes is not None:
            return self.dataframes

        if not self.exists:
            return {}

        try:
            import pandas as pd
            
            # Handle CSV files differently
            file_ext = os.path.splitext(self.filepath)[1].lower()
            if file_ext == '.csv':
                df = pd.read_csv(self.filepath)
                return {'Sheet1': df}  # Return as a single sheet
            
            # Determine the engine based on file extension
            engine = None
            if file_ext in ['.xlsx', '.xlsm']:
                engine = 'openpyxl'
            elif file_ext == '.xls':
                engine = 'xlrd'
            elif file_ext == '.xlsb':
                engine = 'pyxlsb'
            
            # Try opening with the selected engine
            try:
                xl = pd.ExcelFile(self.filepath, engine=engine)
            except Exception:
                # If initial attempt fails, try other engines
                success = False
                for alt_engine in ['openpyxl', 'xlrd', 'pyxlsb']:
                    if alt_engine == engine:  # Skip already tried engine
                        continue
                    try:
                        xl = pd.ExcelFile(self.filepath, engine=alt_engine)
                        success = True
                        break  # Found a working engine
                    except Exception:
                        continue
                
                if not success:
                    raise ValueError(f"Could not determine Excel engine for {os.path.basename(self.filepath)}. Try installing required packages (openpyxl, xlrd, pyxlsb) or specify an engine manually.")
            
            # Process with the working engine
            result = {}
            for sheet in xl.sheet_names:
                result[sheet] = xl.parse(sheet)

            self.dataframes = result
            return result
        except ImportError:
            raise ImportError("pandas library not installed")
        except Exception as e:
            raise ValueError(f"Error parsing Excel file: {str(e)}")

    async def to_dict(self) -> Dict[str, List[Dict[str, Any]]]:
        """Convert Excel file to dictionary representations."""
        if self.data_dict is not None:
            return self.data_dict

        dfs = await self.to_dataframes()
        result = {}
        for sheet, df in dfs.items():
            result[sheet] = df.to_dict(orient="records")

        self.data_dict = result
        return result
    
    async def get_content(self) -> Dict[str, Any]:
        """Get the content of the Excel file."""
        if self.data_dict is None:
            self.data_dict = await self.to_dict()
        return self.data_dict


# Register the message types with the message factory
def register_file_message_types(factory: MessageFactory) -> None:
    """Register all file message types with the given message factory."""
    factory.register(FileMessage)
    factory.register(PDFMessage)
    factory.register(PDFWithOCRMessage)
    factory.register(JSONMessage)
    factory.register(ImageMessage)
    factory.register(ExcelMessage)
    factory.register(FileProcessingEvent)
